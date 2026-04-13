import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from collections import defaultdict

st.title('Generador de Tablas de Operación')

to_file = st.file_uploader('Sube el archivo de Tiempos de Operación', type=['xlsx'])
cc_file = st.file_uploader('Sube el archivo de Cortocircuitos', type=['xlsx'])

if to_file and cc_file:
    if st.button('Procesar Datos'):
        with st.spinner('Procesando...'):
            # 1. Tiempos de operación
            df_top = pd.read_excel(to_file, skiprows=1).dropna(how='all', axis=1)
            
            col_tiempo = next((col for col in df_top.columns if 'Alims.Barra' in str(col)), df_top.columns[-1])
            df_top['Tiempo'] = df_top[col_tiempo]
            df_top['Paño'] = df_top['Paño'].str.replace('AT', 'Respaldo TR')
            
            def get_nombre(row):
                pano, rel, func = str(row['Paño']), str(row['Relé']), str(row['Función'])
                if pano.startswith('Respaldo TR'):
                    if func == '51': return '110kV'
                    if func == '51N': return '12,5 kV'
                    return ''
                if pano.startswith('ET'):
                    if func == '51':
                        nums = re.findall(r'\d+', pano)
                        return f"Barra {nums[0]}" if nums else 'Barra'
                    return ''
                if pano.startswith('EBC'):
                    nums = re.findall(r'\d+', pano)
                    return f"Banco Condensador {nums[0]}" if nums else 'Banco Condensador'
                
                name = rel.split('_')[-1] if len(rel.split('_')) > 4 else ''
                return name
            
            df_top['Nombre'] = df_top.apply(get_nombre, axis=1)
            
            mapa_fallas = {
                '3psc': 'TRIFÁSICO', '2psc': 'BIFÁSICO', '2pgf': 'BIFÁSICO A TIERRA',
                '2pgfR=25': 'BIFÁSICO A TIERRA (R=25)', '2pgfR=50': 'BIFÁSICO A TIERRA (R=50)',
                'spgf R=0': 'MONOFÁSICO', 'spgf R=25': 'MONOFÁSICO (R=25)', 'spgf R=50': 'MONOFÁSICO (R=50)'
            }
            df_top['Falla_Nom'] = df_top['Falla'].map(mapa_fallas).fillna(df_top['Falla'])
            
            tabla_tiempos = pd.pivot_table(
                df_top, values='Tiempo', index='Falla_Nom', columns=['Paño', 'Nombre', 'Función'], aggfunc='first'
            ).fillna('--')
            
            def custom_sort_key(col):
                pano = str(col[0])
                if pano.startswith('Respaldo TR'): grupo = 1
                elif pano.startswith('ET'): grupo = 2
                elif pano.startswith('EBC'): grupo = 4 
                elif re.match(r'^E\d+$', pano): grupo = 3
                else: grupo = 5
                    
                nums = re.findall(r'\d+', pano)
                num = int(nums[0]) if nums else 0
                return (grupo, num, pano, col[2], col[1])
            
            tabla_tiempos = tabla_tiempos[sorted(tabla_tiempos.columns, key=custom_sort_key)]
            
            # 2. Cortocircuitos
            df_cc_raw = pd.read_excel(cc_file)
            df_cc = df_cc_raw.iloc[1:9, 0:5].copy()
            df_cc.columns = ['Falla', 'Fase_MT', 'Fase_AT', '3I0_MT', '3I0_AT']
            
            mapa_cc = {
                'TRIFASICO': 'TRIFÁSICO', 'BIFASICO': 'BIFÁSICO', 'BIFASICO A TIERRA': 'BIFÁSICO A TIERRA',
                'BIFASICO A TIERRA R=25': 'BIFÁSICO A TIERRA (R=25)', 'BIFASICO A TIERRA R=50': 'BIFÁSICO A TIERRA (R=50)',
                'MONOFASICO': 'MONOFÁSICO', 'MONOFASICO R=25': 'MONOFÁSICO (R=25)', 'MONOFASICO R=50': 'MONOFÁSICO (R=50)'
            }
            df_cc['Falla_Nom'] = df_cc['Falla'].map(mapa_cc)
            df_cc = df_cc.dropna(subset=['Falla_Nom']).set_index('Falla_Nom')
            
            # 3. Construir tabla de datos estructurada
            fallas_orden = [
                'TRIFÁSICO', 'BIFÁSICO', 'BIFÁSICO A TIERRA', 'BIFÁSICO A TIERRA (R=25)',
                'BIFÁSICO A TIERRA (R=50)', 'MONOFÁSICO', 'MONOFÁSICO (R=25)', 'MONOFÁSICO (R=50)'
            ]
            
            filas_finales = []
            for falla in fallas_orden:
                fila_cc = {}
                for col in tabla_tiempos.columns:
                    es_at = 'Respaldo TR' in str(col[0])
                    es_fase = str(col[2]) == '51'
                    if falla in df_cc.index:
                        fila_cc[col] = df_cc.loc[falla, 'Fase_AT' if es_at else 'Fase_MT'] if es_fase else df_cc.loc[falla, '3I0_AT' if es_at else '3I0_MT']
                    else:
                        fila_cc[col] = '--'
                filas_finales.append(pd.Series(fila_cc, name=f'Cortocircuito {falla} (kA)'))
                if falla in tabla_tiempos.index:
                    filas_finales.append(tabla_tiempos.loc[falla].rename('Tiempo de operación de COCI en alimentadores'))
            
            tabla_final = pd.DataFrame(filas_finales).fillna('--')
            tabla_final.columns = pd.MultiIndex.from_tuples(tabla_final.columns)
            
            # 4. Construcción Manual en openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = 'Resultados'
            
            color_fondo = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            fuente_negrita = Font(bold=True)
            fuente_azul = Font(color="0000FF", bold=True)
            fuente_naranjo = Font(color="FF8C00", bold=True)
            alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
            borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            op_start_col = 2
            op_end_col = op_start_col + len(tabla_final.columns) - 1
            tp_start_col = op_end_col + 1
            
            ws['A1'] = 'Tipo'
            ws['A4'] = 'COCI en los alimentadores en 12,5 kV'
            
            ws.cell(row=1, column=op_start_col, value='Tiempos de operación')
            ws.merge_cells(start_row=1, start_column=op_start_col, end_row=1, end_column=op_end_col)
            
            feeder_cols = defaultdict(list)
            
            for col_idx, col_tuple in enumerate(tabla_final.columns):
                c = col_idx + op_start_col
                ws.cell(row=2, column=c, value=col_tuple[0])
                ws.cell(row=3, column=c, value=col_tuple[1])
                ws.cell(row=4, column=c, value=col_tuple[2])
                pano = str(col_tuple[0])
                feeder_cols[pano].append(c)
            
            def merge_identical_consecutive(ws, row, start_col, end_col):
                start_merge = start_col
                current_val = ws.cell(row=row, column=start_col).value
                for col in range(start_col + 1, end_col + 1):
                    val = ws.cell(row=row, column=col).value
                    if val != current_val or val == '' or val is None:
                        if col - 1 > start_merge:
                            ws.merge_cells(start_row=row, start_column=start_merge, end_row=row, end_column=col-1)
                        start_merge = col
                        current_val = val
                if end_col > start_merge:
                    ws.merge_cells(start_row=row, start_column=start_merge, end_row=row, end_column=end_col)
            
            merge_identical_consecutive(ws, 2, op_start_col, op_end_col)
            merge_identical_consecutive(ws, 3, op_start_col, op_end_col)
            
            pano_tr = next((p for p in feeder_cols.keys() if 'Respaldo TR' in p), None)
            pano_et = next((p for p in feeder_cols.keys() if p.startswith('ET')), None)
            
            normal_feeders = [p for p in feeder_cols.keys() if p != pano_tr and p != pano_et and not p.startswith('EBC')]
            cap_banks = [p for p in feeder_cols.keys() if p.startswith('EBC')]
            feeders_tp = normal_feeders + cap_banks
            
            nombres_dict = {}
            for col in tabla_tiempos.columns:
                if col[1]: nombres_dict[col[0]] = col[1]
            
            et_name = nombres_dict.get(pano_et, pano_et)
            tr_name = pano_tr
            
            tp_row2 = []
            tp_row4 = []
            
            if pano_tr and pano_et:
                tp_row2.append(f"{tr_name} / {et_name}")
                tp_row4.append(f"{pano_tr.replace('Respaldo ', '')}/{pano_et}")
            
            for f in feeders_tp:
                f_name = nombres_dict.get(f, "")
                texto_row2 = f"{et_name} / {f_name}" if f_name else f"{et_name} / {f}"
                tp_row2.append(texto_row2)
                
                f_acronym = f.replace('EBC', 'BC') if f.startswith('EBC') else f
                tp_row4.append(f"{pano_et}/{f_acronym}")
            
            max_excel_col = tp_start_col + len(tp_row2) - 1
            
            ws.cell(row=1, column=tp_start_col, value='Tiempos de paso ') 
            ws.merge_cells(start_row=1, start_column=tp_start_col, end_row=1, end_column=max_excel_col)
            
            for i in range(len(tp_row2)):
                c = tp_start_col + i
                ws.cell(row=2, column=c, value=tp_row2[i])
                ws.merge_cells(start_row=2, start_column=c, end_row=3, end_column=c)
                ws.cell(row=4, column=c, value=tp_row4[i])
            
            for r_idx, (idx_val, row) in enumerate(tabla_final.iterrows()):
                r = r_idx + 5
                ws.cell(row=r, column=1, value=idx_val)
                for col_idx, val in enumerate(row):
                    ws.cell(row=r, column=col_idx + op_start_col, value=val)
            
            def parse_val(v):
                try:
                    if v is None or v == '--' or str(v).strip() == '': return None
                    return float(str(v).replace(',', '.'))
                except ValueError:
                    return None
            
            observaciones = []
            
            for r in range(5, ws.max_row + 1):
                cell_value = ws.cell(row=r, column=1).value
                
                if cell_value and "Tiempo de operación" in str(cell_value):
                    falla_actual = ws.cell(row=r-1, column=1).value.replace('Cortocircuito ', '').replace(' (kA)', '')
                    
                    mins = {}
                    panos_sin_operar = []
                    
                    for pano, cols in feeder_cols.items():
                        t_vals = [parse_val(ws.cell(row=r, column=c).value) for c in cols]
                        valid_vals = [v for v in t_vals if v is not None]
                        
                        if not valid_vals:
                            panos_sin_operar.append(pano)
                            
                        mins[pano] = min(valid_vals) if valid_vals else np.inf
                        
                    if panos_sin_operar:
                        if len(panos_sin_operar) == 1:
                            texto_panos = panos_sin_operar[0]
                        else:
                            texto_panos = ", ".join(panos_sin_operar[:-1]) + " y " + panos_sin_operar[-1]
                        observaciones.append(f"En {falla_actual}, no opera ninguna protección para: {texto_panos}.")
                        
                    col_offset = 0
                    
                    # Validación y registro para TR - ET
                    if pano_tr and pano_et:
                        diff_tr_et = mins.get(pano_tr, np.inf) - mins.get(pano_et, np.inf)
                        cell_tr = ws.cell(row=r, column=tp_start_col + col_offset)
                        cell_tr.value = diff_tr_et if not np.isinf(diff_tr_et) else "--"
                        
                        if not np.isinf(diff_tr_et) and diff_tr_et < 0.3:
                            observaciones.append(f"En {falla_actual}, no se cumple el tiempo de paso (< 300 ms) entre {pano_tr} y {pano_et}.")
                            
                        col_offset += 1
                        
                    # Validación y registro para ET - Alimentadores/Bancos
                    for idx, f_code in enumerate(feeders_tp):
                        diff = mins.get(pano_et, np.inf) - mins.get(f_code, np.inf)
                        cell_f = ws.cell(row=r, column=tp_start_col + col_offset + idx)
                        cell_f.value = diff if not np.isinf(diff) else "--"
                        
                        if not np.isinf(diff) and diff < 0.3:
                            observaciones.append(f"En {falla_actual}, no se cumple el tiempo de paso (< 300 ms) entre {pano_et} y {f_code}.")
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_excel_col):
                for cell in row:
                    cell.border = borde_delgado
                    if cell.column > 1:
                        cell.alignment = alineacion_centro
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.000'
            
            for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=max_excel_col):
                for cell in row:
                    cell.fill = color_fondo
                    cell.font = fuente_negrita
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.fill = color_fondo
                    cell.font = fuente_negrita
            
            for r in range(5, ws.max_row + 1):
                cell_value = ws.cell(row=r, column=1).value
                if cell_value and "Tiempo de operación" in str(cell_value):
                    for cols in feeder_cols.values():
                        vals = []
                        for c in cols:
                            val = parse_val(ws.cell(row=r, column=c).value)
                            if val is not None:
                                vals.append((val, c))
                        if vals:
                            min_val = min(v[0] for v in vals)
                            for v, c in vals:
                                if v == min_val:
                                    ws.cell(row=r, column=c).font = fuente_azul
                                    break
                    
                    for c in range(tp_start_col, max_excel_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell.value, (int, float)):
                            cell.font = fuente_azul if cell.value > 0.3 else fuente_naranjo
            
            ws.column_dimensions['A'].width = 35
            
            if observaciones:
                fila_obs = ws.max_row + 2
                ws.cell(row=fila_obs, column=1, value="Observaciones:").font = fuente_negrita
                for i, obs in enumerate(observaciones):
                    ws.cell(row=fila_obs + 1 + i, column=1, value=obs)
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success('Documento generado.')
            
            st.download_button(
                label="Descargar Excel",
                data=output,
                file_name="Resultados_Tiempos_Paso.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
